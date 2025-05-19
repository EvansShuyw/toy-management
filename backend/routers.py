from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Body, Response
from starlette.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List
import models
from datetime import datetime
import os
import time
import shutil
import uuid
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
import openpyxl.styles
import os
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
from import_service import import_items

router = APIRouter()

# 配置图片上传目录
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 获取所有货物报价表项目
@router.get("/items/", response_model=List[dict])
def get_items(name: str = None, factory_name: str = None, factory_code: str = None, db: Session = Depends(models.get_db)):
    query = db.query(models.ToyItem)
    if name:
        query = query.filter(models.ToyItem.name.ilike(f"%{name}%"))
    if factory_name:
        query = query.filter(models.ToyItem.factory_name.ilike(f"%{factory_name}%"))
    if factory_code:
        query = query.filter(models.ToyItem.factory_code.ilike(f"%{factory_code}%"))
    items = query.all()
    return [{
        "id": item.id,
        "factory_code": item.factory_code,
        "factory_name": item.factory_name,
        "name": item.name,
        "packaging": item.packaging,
        "packing_quantity": item.packing_quantity,
        "unit_price": item.unit_price,
        "gross_weight": item.gross_weight,
        "net_weight": item.net_weight,
        "outer_box_size": item.outer_box_size,
        "product_size": item.product_size,
        "inner_box": item.inner_box,
        "remarks": item.remarks,
        "image_path": item.image_path,
        "origin_sheet": item.origin_sheet,
        "created_at": item.created_at,
        "updated_at": item.updated_at
    } for item in items]

# 创建新的货物报价表项目
@router.post("/items/")
async def create_item(factory_code: str = Form(...),
                     factory_name: str = Form(...),
                     name: str = Form(...),
                     packaging: str = Form(...),
                     packing_quantity: int = Form(...),
                     unit_price: float = Form(...),
                     gross_weight: float = Form(...),
                     net_weight: float = Form(...),
                     outer_box_size: str = Form(...),
                     product_size: str = Form(...),
                     inner_box: str = Form(...),
                     remarks: str = Form(None),
                     image: UploadFile = File(None),
                     db: Session = Depends(models.get_db)):
    # 处理图片上传
    image_path = None
    if image:
        # 验证文件类型，只允许图片文件
        file_ext = os.path.splitext(image.filename)[1].lower()
        allowed_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp')
        if file_ext not in allowed_extensions:
            raise HTTPException(status_code=400, detail="只允许上传图片文件（JPG、JPEG、PNG、GIF、BMP）")
            
        # 保存文件前验证文件内容是否为有效图片
        try:
            # 保存文件内容到内存中进行验证
            contents = await image.read()
            try:
                # 尝试用PIL打开图片验证格式
                img = PILImage.open(BytesIO(contents))
                img.verify()  # 验证图片完整性
                # 重置文件指针，以便后续保存文件
                await image.seek(0)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"无效的图片文件: {str(e)}")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"文件验证失败: {str(e)}")
            
        file_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}{file_ext}"
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        try:
            file_path = os.path.join(UPLOAD_DIR, file_name)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(image.file, buffer)
            # 确保文件路径正确
            image_path = os.path.relpath(file_path, os.path.dirname(__file__)).replace('\\', '/')
            # 设置为相对URL路径以便前端访问
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")
    
    
    # 创建数据库记录
    db_item = models.ToyItem(
        factory_code=factory_code,
        factory_name=factory_name,
        name=name,
        packaging=packaging,
        packing_quantity=packing_quantity,
        unit_price=unit_price,
        gross_weight=gross_weight,
        net_weight=net_weight,
        outer_box_size=outer_box_size,
        product_size=product_size,
        inner_box=inner_box,
        remarks=remarks,
        image_path=image_path
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

# 更新货物报价表项目
@router.put("/items/{item_id}")
async def update_item(item_id: int,
                     factory_code: str = Form(...),
                     factory_name: str = Form(...),
                     name: str = Form(...),
                     packaging: str = Form(...),
                     packing_quantity: int = Form(...),
                     unit_price: float = Form(...),
                     gross_weight: float = Form(...),
                     net_weight: float = Form(...),
                     outer_box_size: str = Form(...),
                     product_size: str = Form(...),
                     inner_box: str = Form(...),
                     remarks: str = Form(None),
                     image: UploadFile = File(None),
                     db: Session = Depends(models.get_db)):
    # 查找要更新的记录
    db_item = db.query(models.ToyItem).filter(models.ToyItem.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # 处理图片上传
    image_path = db_item.image_path
    if image:
        # 验证文件类型，只允许图片文件
        file_ext = os.path.splitext(image.filename)[1].lower()
        allowed_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp')
        if file_ext not in allowed_extensions:
            raise HTTPException(status_code=400, detail="只允许上传图片文件（JPG、JPEG、PNG、GIF、BMP）")
            
        # 保存文件前验证文件内容是否为有效图片
        try:
            # 保存文件内容到内存中进行验证
            contents = await image.read()
            try:
                # 尝试用PIL打开图片验证格式
                img = PILImage.open(BytesIO(contents))
                img.verify()  # 验证图片完整性
                # 重置文件指针，以便后续保存文件
                await image.seek(0)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"无效的图片文件: {str(e)}")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"文件验证失败: {str(e)}")
            
        # 删除旧图片
        if db_item.image_path:
            old_file_path = os.path.join(os.path.dirname(__file__), db_item.image_path.lstrip('/'))
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
        
        # 保存新图片
        file_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}{file_ext}"
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        try:
            file_path = os.path.join(UPLOAD_DIR, file_name)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(image.file, buffer)
            # 设置为相对URL路径以便前端访问
            image_path = f"uploads/{file_name}"
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"文件保存失败: {str(e)}")
    
    # 更新记录
    db_item.factory_code = factory_code
    db_item.factory_name = factory_name
    db_item.name = name
    db_item.packaging = packaging
    db_item.packing_quantity = packing_quantity
    db_item.unit_price = unit_price
    db_item.gross_weight = gross_weight
    db_item.net_weight = net_weight
    db_item.outer_box_size = outer_box_size
    db_item.product_size = product_size
    db_item.inner_box = inner_box
    db_item.remarks = remarks
    db_item.image_path = image_path
    db_item.updated_at = datetime.now()
    
    db.commit()
    db.refresh(db_item)
    return db_item

# 删除货物报价表项目
@router.delete("/items/{item_id}")
def delete_item(item_id: int, db: Session = Depends(models.get_db)):
    db_item = db.query(models.ToyItem).filter(models.ToyItem.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # 删除关联的图片文件
    if db_item.image_path:
        file_path = os.path.join(UPLOAD_DIR, os.path.basename(db_item.image_path))
        if os.path.exists(file_path):
            os.remove(file_path)
    
    db.delete(db_item)
    db.commit()
    return {"message": "Item deleted successfully"}

import asyncio
from concurrent.futures import ThreadPoolExecutor
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 导出选中的货物报价表为Excel
@router.post("/items/export")
async def export_items(request: dict = Body(...), db: Session = Depends(models.get_db)):
    item_ids = request.get("item_ids", [])
    # 获取选中的项目
    items = db.query(models.ToyItem).filter(models.ToyItem.id.in_(item_ids)).all()
    if not items:
        raise HTTPException(status_code=400, detail="No items found for export")
    
    # 创建内存缓存，避免重复处理相同的图片
    image_cache = {}
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "货物报价表"
    
    # 记录开始时间
    start_time = time.time()
    
    # 添加表头
    headers = ["图片", "货号", "厂名", "品名", "包装", "装箱量PCS", "单价", "毛重KG", "净重KG", "外箱规格CM", "产品规格", "内箱", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # 设置表头单元格样式：水平居中、自动换行
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 设置列宽 - 按照要求设置
    # 图片列宽设为20px
    ws.column_dimensions['A'].width = 20
    # 品名列宽设为30px
    ws.column_dimensions['D'].width = 30
    # 其他列设置合适的宽度
    for col_idx in range(2, len(headers) + 1):
        if col_idx != 4:  # 跳过品名列，因为已经单独设置了
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = 15
    
    # 添加数据
    for row, item in enumerate(items, 2):
        # 设置行高为100px（Excel中的行高单位约为0.75pt = 1px）
        ws.row_dimensions[row].height = 75  # 100px 约等于 75pt
        
        # 添加数据并设置单元格样式
        for col_idx, value in enumerate([
            None,  # 图片列稍后处理
            item.factory_code,
            item.factory_name,
            item.name,
            item.packaging,
            item.packing_quantity,
            item.unit_price,
            item.gross_weight,
            item.net_weight,
            item.outer_box_size,
            item.product_size,
            item.inner_box,
            item.remarks
        ], 1):
            if col_idx > 1:  # 跳过图片列
                cell = ws.cell(row=row, column=col_idx, value=value)
                # 设置所有文本单元格水平居中并启用自动换行
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 处理图片
        if item.image_path and os.path.exists(item.image_path):
            # 检查缓存中是否已有相同的图片
            image_hash = hash(open(item.image_path, 'rb').read())
            if image_hash in image_cache:
                # 使用缓存的图片对象
                img = image_cache[image_hash]
                img.anchor = OneCellAnchor(
                    _from=AnchorMarker(col=0, colOff=0, row=row-1, rowOff=0),
                    ext=img.anchor.ext
                )
                ws.add_image(img)
            else:
                # 使用PIL打开并转换图片
                pil_image = PILImage.open(item.image_path)
                # 转换图片为RGB模式（如果是RGBA）
                if pil_image.mode in ('RGBA', 'LA'):
                    background = PILImage.new('RGB', pil_image.size, (255, 255, 255))
                    background.paste(pil_image, mask=pil_image.split()[-1])
                    pil_image = background
                
                # 获取原始图片尺寸
                img_width, img_height = pil_image.size
                
                # 计算单元格的实际尺寸（像素）
                cell_width = ws.column_dimensions['A'].width * 7  # 列宽转换为像素
                cell_height = ws.row_dimensions[row].height * 0.75  # 行高转换为像素
                
                # 保存为临时的BytesIO对象，使用优化的设置
                img_byte_arr = BytesIO()
                # 使用JPEG格式和适当的质量设置，在文件大小和质量之间取得平衡
                pil_image.save(img_byte_arr, format='JPEG', optimize=True, quality=85)
                img_byte_arr.seek(0)
                
                # 创建openpyxl图片对象
                img = Image(img_byte_arr)
                
                # 导入必要的类
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.utils.units import pixels_to_EMU, points_to_pixels
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                
                # 获取单元格的实际尺寸（以像素为单位）
                cell_width = points_to_pixels(ws.column_dimensions['A'].width * 7)  # 列宽转换为像素
                cell_height = points_to_pixels(ws.row_dimensions[row].height)  # 行高转换为像素
                
                # 计算图片的缩放比例，保持宽高比
                padding = 4  # 每边2像素的内边距
                max_width = cell_width - padding
                max_height = cell_height - padding
                scale = min(max_width/img_width, max_height/img_height)
                
                # 计算缩放后的尺寸
                new_width = int(img_width * scale)
                new_height = int(img_height * scale)
                
                # 计算图片在单元格中的偏移量，使其居中
                col_offset = pixels_to_EMU((cell_width - new_width) // 2)
                row_offset = pixels_to_EMU((cell_height - new_height) // 2)
                
                # 创建单元格锚点标记，使用计算出的偏移量
                marker = AnchorMarker(col=0, colOff=col_offset, row=row-1, rowOff=row_offset)
                
                # 创建图片尺寸对象（EMU单位）
                size = XDRPositiveSize2D(pixels_to_EMU(new_width), pixels_to_EMU(new_height))
                
                # 创建单元格锚点
                anchor = OneCellAnchor(_from=marker, ext=size)
                img.anchor = anchor
                
                # 设置图片为单元格内容，并禁止编辑
                img.anchor.editAs = 'oneCell'
                
                # 将图片添加到工作表
                ws.add_image(img)
                
                # 将处理后的图片对象存入缓存
                image_cache[image_hash] = img
            
            # 设置图片单元格的对齐方式
            cell = ws.cell(row=row, column=1)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    
    # 记录处理完成时间
    process_time = time.time() - start_time
    logger.info(f"数据处理完成，耗时：{process_time:.2f}秒")
    
    # 保存Excel文件
    save_start_time = time.time()
    file_name = f"货物报价表_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    os.makedirs("exports", exist_ok=True)
    file_path = os.path.join("exports", file_name)
    wb.save(file_path)
    
    # 记录保存完成时间
    save_time = time.time() - save_start_time
    logger.info(f"文件保存完成，耗时：{save_time:.2f}秒")
    
    # 返回文件并在发送后删除
    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(lambda: os.unlink(file_path))
    )

# 导入Excel数据
@router.post("/items/import")
async def import_excel(file: UploadFile = File(...), factory_name: str = Form(...), db: Session = Depends(models.get_db)):
    # 调用import_service中的import_items函数
    return await import_items(file=file, factory_name=factory_name, db=db)

# 导入Excel数据路由已统一，使用唯一的导入方法

# 获取Excel导入模板
@router.get("/items/import-template")
async def get_import_template():
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "货物导入模板"
    
    # 添加表头
    headers = ["图片", "货号", "厂名", "品名", "包装", "装箱量PCS", "单价", "毛重KG", "净重KG", "外箱规格CM", "产品规格", "内箱", "备注"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # 设置列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 15
    
    # 特别设置图片列的宽度
    ws.column_dimensions['A'].width = 20
    
    # 保存到临时文件
    template_path = "exports/import_template.xlsx"
    os.makedirs("exports", exist_ok=True)
    wb.save(template_path)
    
    # 返回文件
    return FileResponse(
        path=template_path,
        filename="货物导入模板.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(lambda: os.unlink(template_path))
    )