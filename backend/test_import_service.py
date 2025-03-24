import pytest
from fastapi import UploadFile, HTTPException
from sqlalchemy.orm import Session
from unittest.mock import MagicMock, patch
from datetime import datetime
import os
import shutil
import uuid
from openpyxl import Workbook
from PIL import Image as PILImage
from io import BytesIO

from import_service import import_items
from models import ToyItem

@pytest.fixture
def mock_db():
    db = MagicMock(spec=Session)
    return db

@pytest.fixture
def mock_upload_file(tmp_path):
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    headers = ["图片", "货号", "厂名", "品名", "包装", "装箱量PCS", "单价", "毛重KG", "净重KG", "外箱规格CM", "产品规格", "内箱", "备注"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    wb.save(file_path)
    return UploadFile(filename="test.xlsx", file=open(file_path, "rb"))

@patch('import_service.os.makedirs')
@patch('import_service.os.path.exists', return_value=True)
@patch('import_service.os.remove')
@patch('import_service.shutil.copyfileobj')
@patch('import_service.load_workbook')
@patch('import_service.PILImage.open')
@patch('import_service.uuid.uuid4')
@patch('import_service.datetime')
async def test_import_items(mock_datetime, mock_uuid, mock_pil, mock_load_workbook, mock_copyfileobj, mock_remove, mock_exists, mock_makedirs, mock_db, mock_upload_file):
    # Mock datetime
    mock_datetime.now.return_value.strftime.return_value = "20230101000000"
    # Mock uuid
    mock_uuid.return_value.hex = "testuuid"
    # Mock PILImage
    mock_image = MagicMock()
    mock_image.mode = "RGB"
    mock_pil.return_value = mock_image
    # Mock workbook
    mock_workbook = MagicMock()
    mock_sheet = MagicMock()
    mock_sheet._images = []
    mock_sheet.iter_rows.return_value = [
        [MagicMock(value="test_code"), MagicMock(value="test_name"), MagicMock(value="test_packaging"), MagicMock(value=10)]
    ]
    mock_workbook.active = mock_sheet
    mock_load_workbook.return_value = mock_workbook

    # Test
    result = await import_items(file=mock_upload_file, factory_name="test_factory", db=mock_db)

    # Assert
    assert result["imported_count"] == 1
    mock_db.add.assert_called_once()
    mock_db.commit.assert_called_once()

async def test_import_items_invalid_file_type(mock_db):
    mock_file = MagicMock(spec=UploadFile)
    mock_file.filename = "test.txt"
    with pytest.raises(HTTPException) as exc_info:
        await import_items(file=mock_file, factory_name="test_factory", db=mock_db)
    assert exc_info.value.status_code == 400
    assert "只支持Excel文件格式" in exc_info.value.detail

async def test_import_items_with_images(mock_db):
    # Setup mock upload file with images
    file_path = "d:\\toy-manage-system\\backend\\exports\\货物导入模板 (2).xlsx"
    mock_file = UploadFile(filename="货物导入模板 (2).xlsx", file=open(file_path, "rb"))

    # Mock necessary components
    with patch('import_service.load_workbook') as mock_load_workbook, \
         patch('import_service.PILImage.open') as mock_pil, \
         patch('import_service.uuid.uuid4') as mock_uuid, \
         patch('import_service.datetime') as mock_datetime, \
         patch('import_service.os.makedirs'), \
         patch('import_service.os.path.exists', return_value=True):

        # Mock UUID
        mock_uuid.return_value.hex = "testuuid"
        # Mock datetime
        mock_datetime.now.return_value.strftime.return_value = "20230101000000"
        # Mock PILImage
        mock_image = MagicMock()
        mock_image.mode = "RGB"
        mock_pil.return_value = mock_image

        # Execute import
        result = await import_items(file=mock_file, factory_name="test_factory", db=mock_db)

        # Verify results
        assert result["imported_count"] > 0
        mock_db.add.assert_called()
        mock_db.commit.assert_called()

async def test_import_items_missing_required_columns(mock_db, mock_upload_file):
    with patch('import_service.load_workbook') as mock_load_workbook:
        mock_workbook = MagicMock()
        mock_sheet = MagicMock()
        mock_sheet._images = []
        mock_sheet.iter_rows.return_value = []
        mock_workbook.active = mock_sheet
        mock_load_workbook.return_value = mock_workbook
        with pytest.raises(HTTPException) as exc_info:
            await import_items(file=mock_upload_file, factory_name="test_factory", db=mock_db)
        assert exc_info.value.status_code == 400
        assert "缺少必要的列" in exc_info.value.detail