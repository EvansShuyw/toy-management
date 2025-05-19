from fastapi import UploadFile, File, Form, HTTPException, Depends
from sqlalchemy.orm import Session
import models
from datetime import datetime
import os
import shutil
import uuid
from openpyxl import load_workbook
from PIL import Image as PILImage
from io import BytesIO
import zipfile
import re
import asyncio
from concurrent.futures import ThreadPoolExecutor
import time
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")

async def import_items(
    file: UploadFile = File(...),
    factory_name: str = Form(...),
    db: Session = Depends(models.get_db),
    batch_size: int = 50,  # 批量提交数据库的大小
    max_workers: int = 8,  # 并行处理图片的工作线程数
    max_image_size: int = 3500,  # 图片最大尺寸（宽或高）
    image_quality: int = 100,  # 图片压缩质量
    timeout_per_image: int = 10  # 每张图片处理的超时时间（秒）
):
    # 检查文件类型
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="只支持Excel文件格式(.xlsx, .xls)")

    # 创建上传目录
    os.makedirs(UPLOAD_DIR, exist_ok=True)

    # 保存上传的文件到临时位置
    start_time = time.time()
    temp_file_path = f"uploads/temp_{file.filename}"
    with open(temp_file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    logger.info(f"文件保存耗时: {time.time() - start_time:.2f}秒")

    try:
        # 使用新方法处理Excel文件
        # 1. 将.xlsx文件当作zip文件打开，直接提取其中的图片
        images = []
        image_extraction_start = time.time()
        if file.filename.endswith('.xlsx'):
            try:
                # 尝试从xlsx文件中提取图片
                with zipfile.ZipFile(temp_file_path, 'r') as zip_ref:
                    # 查找所有图片文件
                    media_files = [item for item in zip_ref.namelist()
                                  if item.startswith('xl/media/') and
                                  item.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]

                    # 并行提取图片数据
                    for item in media_files:
                        # 提取图片数据
                        image_data = zip_ref.read(item)
                        images.append(image_data)
                        logger.info(f"从Excel中提取图片: {item}")

                    logger.info(f"共提取了 {len(images)} 张图片")
            except Exception as e:
                logger.error(f"提取Excel图片时出错: {str(e)}")
                # 继续处理，即使没有图片

        logger.info(f"图片提取耗时: {time.time() - image_extraction_start:.2f}秒")

        # 2. 使用openpyxl读取所有sheet数据
        workbook = load_workbook(temp_file_path)
        total_imported = 0

        # 遍历所有工作表
        for sheet_idx, sheet in enumerate(workbook.worksheets, 1):
            # 跳过空表
            if sheet.max_row < 2:
                continue

            # 获取表头（第一行）
            headers = [cell.value for cell in sheet[1]]
            logger.info(f"正在处理工作表 {sheet.title}（第{sheet_idx}个），共{sheet.max_row-1}行数据")

            # 定义字段映射（Excel列名 -> 数据库字段名）
            field_mapping = {
                "图片": "image_path",
                "货号": "factory_code",
                "厂名": "factory_name",
                "品名": "name",
                "包装": "packaging",
                "装箱量PCS": "packing_quantity",
                "单价": "unit_price",
                "毛重KG": "gross_weight",
                "净重KG": "net_weight",
                "外箱规格CM": "outer_box_size",
                "产品规格": "product_size",
                "内箱": "inner_box",
                "备注": "remarks"
            }

            # 找出每个字段在Excel中的列索引
            field_indices = {}
            for i, header in enumerate(headers):
                if header in field_mapping:
                    field_indices[field_mapping[header]] = i

            # 检查必填字段是否存在
            required_fields = ["factory_code", "name", "packaging", "packing_quantity"]
            missing_fields = [field for field in required_fields if field not in field_indices]
            if missing_fields:
                missing_headers = [key for key, value in field_mapping.items() if value in missing_fields]
                raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_headers)}")

            # 创建内存缓存，避免重复处理相同的图片数据
            image_cache = {}

            # 定义图片处理函数
            async def process_image(image_data, row_idx):
                # 检查缓存中是否已有相同的图片数据
                image_hash = hash(image_data)
                if image_hash in image_cache:
                    logger.info(f"使用缓存的图片处理结果: {row_idx}")
                    return image_cache[image_hash]

                try:
                    # 使用超时机制处理图片
                    try:
                        # 从图片数据创建PIL Image对象
                        img = PILImage.open(BytesIO(image_data))

                        # 如果是RGBA模式，转换为RGB
                        if img.mode in ('RGBA', 'LA'):
                            background = PILImage.new('RGB', img.size, (255, 255, 255))
                            background.paste(img, mask=img.split()[-1])
                            img = background

                        # 调整图片大小以减少处理时间和文件大小
                        width, height = img.size
                        if width > max_image_size or height > max_image_size:
                            # 保持宽高比例
                            if width > height:
                                new_width = max_image_size
                                new_height = int(height * (max_image_size / width))
                            else:
                                new_height = max_image_size
                                new_width = int(width * (max_image_size / height))
                            # 使用LANCZOS重采样算法获得更好的质量和速度平衡
                            img = img.resize((new_width, new_height), PILImage.LANCZOS)
                            logger.info(f"调整图片大小: {width}x{height} -> {new_width}x{new_height}")
                    except Exception as e:
                        logger.error(f"无效的图片格式: {str(e)}")
                        return None  # 跳过无效的图片

                    # 生成唯一的文件名
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    unique_id = uuid.uuid4().hex[:8]
                    # 使用jpg格式替代png，可以大幅减小文件大小
                    file_name = f"{timestamp}_{unique_id}_{row_idx}.jpg"
                    file_path = os.path.abspath(os.path.join(UPLOAD_DIR, file_name))

                    # 确保文件路径在uploads目录内
                    if not os.path.commonpath([file_path, os.path.abspath(UPLOAD_DIR)]) == os.path.abspath(UPLOAD_DIR):
                        raise ValueError("无效的文件路径")

                    # 保存为JPG格式 - 使用异步IO操作和超时控制
                    def save_image():
                        # 使用优化参数保存图片
                        img.save(file_path, 'JPEG',
                                 optimize=True,
                                 quality=image_quality,  # 降低质量以减小文件大小
                                 progressive=True)  # 使用渐进式JPEG提高加载体验
                        return file_path

                    # 在线程池中执行IO密集型操作，并添加超时控制
                    loop = asyncio.get_event_loop()
                    try:
                        # 使用超时机制避免单个图片处理时间过长
                        saved_path = await asyncio.wait_for(
                            loop.run_in_executor(None, save_image),
                            timeout=timeout_per_image
                        )

                        logger.info(f'成功保存图片到：{saved_path}')

                        # 设置为相对URL路径以便前端访问
                        image_path = f"uploads/{file_name}"

                        # 将结果存入缓存
                        image_cache[image_hash] = image_path

                        return image_path
                    except asyncio.TimeoutError:
                        logger.warning(f"处理第 {row_idx} 行图片超时，跳过处理")
                        return None
                except Exception as e:
                    logger.error(f"处理第 {row_idx} 行图片时出错: {str(e)}")
                    return None

            # 预处理图片 - 并行处理所有图片
            image_paths = {}
            if images:
                logger.info(f"开始并行处理 {len(images)} 张图片")
                image_processing_start = time.time()

                # 检测重复图片数据，减少处理量
                unique_images = {}
                for img_index, image_data in enumerate(images):
                    image_hash = hash(image_data)
                    if image_hash not in unique_images:
                        unique_images[image_hash] = [img_index]
                    else:
                        unique_images[image_hash].append(img_index)

                logger.info(f"检测到 {len(images)} 张图片中有 {len(unique_images)} 张唯一图片")

                # 创建线程池执行器
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    # 创建图片处理任务 - 只处理唯一的图片
                    tasks = []
                    for image_hash, indices in unique_images.items():
                        # 使用第一个索引作为处理标识
                        img_index = indices[0]
                        task = process_image(images[img_index], img_index)
                        tasks.append((task, indices))

                    # 等待所有图片处理完成，使用as_completed获取先完成的结果
                    completed_tasks = 0
                    total_tasks = len(tasks)

                    # 使用semaphore限制并发任务数量，避免内存过载
                    semaphore = asyncio.Semaphore(max_workers * 2)

                    async def process_with_semaphore(task_info):
                        task, indices = task_info
                        async with semaphore:
                            result = await task
                            return result, indices

                    # 创建所有任务
                    all_tasks = [process_with_semaphore(task_info) for task_info in tasks]

                    # 处理完成的任务
                    for future in asyncio.as_completed(all_tasks):
                        result, indices = await future
                        completed_tasks += 1

                        # 更新进度
                        if completed_tasks % 5 == 0 or completed_tasks == total_tasks:
                            logger.info(f"图片处理进度: {completed_tasks}/{total_tasks}")

                        if result:  # 只存储成功处理的图片路径
                            # 将结果应用到所有使用相同图片的索引
                            for img_index in indices:
                                image_paths[img_index] = result

                logger.info(f"图片处理完成，共处理 {len(image_paths)} 张图片，耗时: {time.time() - image_processing_start:.2f}秒")

            # 导入数据
            data_import_start = time.time()
            imported_count = 0
            batch_items = []  # 用于批量提交的项目列表

            # 计算全局行号（跨sheet累计）
            global_row_offset = total_imported
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # 从第二行开始（跳过表头）
                # 计算全局行索引（用于图片匹配）
                global_row_index = global_row_offset + (row_idx - 2)
                try:
                    # 获取图片路径（如果有）
                    image_path = None
                    if "image_path" in field_indices and image_paths:
                        # 计算当前行对应的图片索引
                        img_index = row_idx - 2  # 从0开始
                        # 如果有对应的图片，使用预处理的路径
                        image_path = image_paths.get(img_index)

                    # 创建新记录
                    new_item = models.ToyItem(
                        factory_code=row[field_indices.get("factory_code")].value if "factory_code" in field_indices else None,
                        factory_name=row[field_indices.get("factory_name")].value if "factory_name" in field_indices else factory_name,
                        name=row[field_indices.get("name")].value if "name" in field_indices else None,
                        packaging=row[field_indices.get("packaging")].value if "packaging" in field_indices else None,
                        packing_quantity=int(row[field_indices.get("packing_quantity")].value) if "packing_quantity" in field_indices and row[field_indices.get("packing_quantity")].value is not None else 0,
                        unit_price=float(row[field_indices.get("unit_price")].value) if "unit_price" in field_indices and row[field_indices.get("unit_price")].value is not None else 0.0,
                        gross_weight=float(row[field_indices.get("gross_weight")].value) if "gross_weight" in field_indices and row[field_indices.get("gross_weight")].value is not None else 0.0,
                        net_weight=float(row[field_indices.get("net_weight")].value) if "net_weight" in field_indices and row[field_indices.get("net_weight")].value is not None else 0.0,
                        outer_box_size=row[field_indices.get("outer_box_size")].value if "outer_box_size" in field_indices else None,
                        product_size=row[field_indices.get("product_size")].value if "product_size" in field_indices else None,
                        inner_box=row[field_indices.get("inner_box")].value if "inner_box" in field_indices else None,
                        remarks=row[field_indices.get("remarks")].value if "remarks" in field_indices else None,
                        image_path=image_path,
                        origin_sheet=sheet.title  # 记录来源工作表
                    )

                    # 如果提供了厂名但Excel中没有厂名字段，使用表单提供的厂名
                    if not new_item.factory_name:
                        new_item.factory_name = factory_name

                    # 检查必填字段
                    missing_fields = []
                    if not new_item.factory_code:
                        missing_fields.append("货号")
                    if not new_item.name:
                        missing_fields.append("品名")
                    if not new_item.packaging:
                        missing_fields.append("包装")

                    if missing_fields:
                        logger.warning(f"导入第 {row_idx} 行时缺少必填字段: {', '.join(missing_fields)}")
                        continue

                    try:
                        # 添加到批处理列表
                        batch_items.append(new_item)
                        imported_count += 1

                        # 当达到批处理大小时，批量提交到数据库
                        if len(batch_items) >= batch_size:
                            db.add_all(batch_items)
                            db.flush()  # 刷新但不提交
                            batch_items = []  # 清空批处理列表
                            logger.info(f"已批量处理 {imported_count} 条记录")

                    except Exception as e:
                        logger.error(f"导入第 {row_idx} 行时数据库操作失败: {str(e)}")
                        continue
                except Exception as e:
                    # 记录错误但继续处理其他行
                    logger.error(f"导入第 {row_idx} 行时出错: {str(e)}")
                    continue

            # 处理剩余的批次
            if batch_items:
                db.add_all(batch_items)

            # 提交事务
            commit_start = time.time()
            db.commit()
            logger.info(f"数据库提交耗时: {time.time() - commit_start:.2f}秒")
            logger.info(f"数据导入耗时: {time.time() - data_import_start:.2f}秒")

            # 计算总耗时
            total_time = time.time() - start_time
            logger.info(f"总耗时: {total_time:.2f}秒")

            # 累计总导入数量
            total_imported += imported_count

        # 返回导入结果
        return {
            "imported_count": total_imported,
            "message": f"成功导入{len(workbook.worksheets)}个工作表",
            "total_time": f"{total_time:.2f}秒"
        }

    except Exception as e:
        # 回滚事务
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")

    finally:
        # 清理临时文件
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)